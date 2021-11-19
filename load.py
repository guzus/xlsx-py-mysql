from database import Database
import logging
import openpyxl
import settings
import sys


def load_worksheet():
    wb = openpyxl.load_workbook(settings.XLSX_FILE_PATH, read_only=True, data_only=True)
    logging.info(f"Found {wb.sheetnames} from [{settings.XLSX_FILE_PATH}]")

    ws = wb[settings.XLSX_FILE_SHEET]
    logging.info(f"Found {ws}")

    return ws


def refine_data(data):
    return data


def load_data(db, ws):
    for xlsx_row in range(settings.XLSX_ROW_RANGE[0], settings.XLSX_ROW_RANGE[1] + 1):
        data = {}
        for xlsx_column, table_column in settings.XLSX_COLUMN_TO_DB_TABLE_COLUMN.items():
            key = table_column
            value = ws[f"{xlsx_column}{xlsx_row}"].value
            data[key] = value

        db.insert(column=refine_data(data))
    db.draw_result()


if __name__ == "__main__":
    if "log" in sys.argv:
        logging.getLogger().setLevel(logging.INFO)
    db = Database()
    ws = load_worksheet()
    load_data(db, ws)
